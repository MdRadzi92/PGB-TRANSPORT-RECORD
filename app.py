import streamlit as st
import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook

st.set_page_config(page_title="Transport Record", layout="wide")

EXCEL_PATH = os.path.join("data", "transport_demo.xlsx")

# ---------------- Helpers ----------------
@st.cache_data(ttl=5)
def load_sheets():
    users = pd.read_excel(EXCEL_PATH, sheet_name="Users")
    vehicles = pd.read_excel(EXCEL_PATH, sheet_name="Vehicles")
    usage = pd.read_excel(EXCEL_PATH, sheet_name="UsageLog")
    settings = pd.read_excel(EXCEL_PATH, sheet_name="Settings")
    # Coerce types
    if "Date" in usage.columns:
        usage["Date"] = pd.to_datetime(usage["Date"], errors="coerce").dt.date
    for col in ["OdoStart", "OdoEnd", "Distance"]:
        if col in usage.columns:
            usage[col] = pd.to_numeric(usage[col], errors="coerce")
    return users, vehicles, usage, settings

def save_sheet(df: pd.DataFrame, sheet_name: str):
    wb = load_workbook(EXCEL_PATH)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def get_setting(key, default=None):
    _, _, _, settings = load_sheets()
    row = settings[settings["Key"] == key]
    if row.empty:
        return default
    val = str(row.iloc[0]["Value"]).strip()
    if val.upper() in ["TRUE", "FALSE"]:
        return val.upper() == "TRUE"
    try:
        return float(val)
    except:
        return val

def authenticate(username, password):
    users, *_ = load_sheets()
    row = users[(users["Username"]==username) & (users["Password"]==password)]
    if row.empty:
        return None
    r = row.iloc[0].to_dict()
    return {"username": r["Username"], "role": r["Role"], "fullname": r.get("FullName",""), "company": r.get("Company","")}

def compute_distance(start, end):
    if pd.isna(start) or pd.isna(end): return None
    return float(end) - float(start)

def flag_daily_anomaly(distance):
    limit = get_setting("DAILY_TRIP_LIMIT", 1000)
    if distance is None: return ""
    return "DAILY_HIGH" if distance > float(limit) else ""

def monthly_vehicle_distance(usage_df):
    if usage_df.empty:
        return pd.DataFrame(columns=["VehicleID","Month","MonthlyDistance"])
    df = usage_df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Month"] = df["Date"].dt.to_period("M").astype(str)
    grp = df.groupby(["VehicleID","Month"])["Distance"].sum(min_count=1).reset_index(name="MonthlyDistance")
    return grp

def flag_monthly_anomalies(monthly_df):
    jump = float(get_setting("MONTHLY_HIGH_JUMP", 4000))
    monthly_df["MonthlyFlag"] = monthly_df["MonthlyDistance"].apply(lambda x: "MONTHLY_HIGH" if pd.notna(x) and x>jump else "")
    return monthly_df

def service_alerts(vehicles):
    interval = float(get_setting("SERVICE_INTERVAL_KM", 10000))
    v = vehicles.copy()
    v["NeedService"] = (v["Odometer"] - v["LastServiceOdo"]) >= interval
    return v[v["NeedService"]]

# ---------------- Login ----------------
if "auth" not in st.session_state:
    st.session_state.auth = None

with st.sidebar:
    st.title("🔐 Login")
    if st.session_state.auth is None:
        u = st.text_input("Username")
        p = st.text_input("Password", type="password")
        if st.button("Sign in"):
            user = authenticate(u, p)
            if user:
                st.session_state.auth = user
                st.success(f"Welcome {user['fullname'] or user['username']}!")
                st.rerun()
            else:
                st.error("Invalid credentials")
    else:
        st.write(f"Logged in as **{st.session_state.auth['username']}** ({st.session_state.auth['role']})")
        if st.button("Logout"):
            st.session_state.auth = None
            st.rerun()

if st.session_state.auth is None:
    st.info("Please login to continue.")
    st.stop()

auth = st.session_state.auth
role = auth["role"]

# ---------------- Data ----------------
users, vehicles, usage, settings = load_sheets()
if not usage.empty:
    usage["Distance"] = usage.apply(lambda r: compute_distance(r.get("OdoStart"), r.get("OdoEnd")), axis=1)
    usage["AnomalyFlag"] = usage.apply(lambda r: flag_daily_anomaly(r["Distance"]), axis=1)

monthly = flag_monthly_anomalies(monthly_vehicle_distance(usage))

# ---------------- Sidebar Nav ----------------
st.sidebar.markdown("---")
section = st.sidebar.radio("Navigate", ["Dashboard", "Vehicles", "Usage", "Reports"] + (["Admin"] if role=="Admin" else []))

# ---------------- Dashboard ----------------
if section == "Dashboard":
    st.title("🚘 Transport Dashboard")
    col1, col2, col3 = st.columns(3)
    with col1:
        total_vehicles = len(vehicles)
        st.metric("Total Vehicles", total_vehicles)
    with col2:
        available = (vehicles["Status"].str.lower()=="available").sum()
        st.metric("Available Now", int(available))
    with col3:
        in_use = total_vehicles - int(available)
        st.metric("In Use", int(in_use))

    st.subheader("Availability by Company")
    avail = vehicles.copy()
    avail["isAvailable"] = avail["Status"].str.lower()=="available"
    summary = avail.groupby("Company")["isAvailable"].agg(["sum","count"]).reset_index().rename(columns={"sum":"Available","count":"Total"})
    summary["InUse"] = summary["Total"] - summary["Available"]
    st.dataframe(summary, use_container_width=True)

    st.subheader("Monthly Mileage by Vehicle")
    if not monthly.empty:
        pivot = monthly.pivot(index="Month", columns="VehicleID", values="MonthlyDistance").fillna(0.0)
        st.line_chart(pivot)
        flags = monthly[monthly["MonthlyFlag"] != ""]
        if not flags.empty:
            st.warning("Monthly anomalies detected:")
            st.dataframe(flags, use_container_width=True)
    else:
        st.info("No usage data yet.")

    # Service alerts
    st.subheader("🔧 Service Alerts")
    alerts = service_alerts(vehicles)
    if not alerts.empty:
        st.error("Vehicles needing service:")
        st.dataframe(alerts[["VehicleID","PlateNo","Company","Odometer","LastServiceOdo","LastServiceDate"]], use_container_width=True)
    else:
        st.success("No vehicles require service.")

# ---------------- Vehicles ----------------
elif section == "Vehicles":
    st.title("🚗 Vehicles")
    st.dataframe(vehicles, use_container_width=True)
    st.caption("Users are view-only. Admins: use Admin tab to add/update vehicles.")

# ---------------- Usage ----------------
elif section == "Usage":
    st.title("📝 Usage Log")
    st.dataframe(usage.sort_values(by="Date", ascending=False), use_container_width=True)
    st.caption("Users are view-only. Admins: use Admin tab to add usage.")

# ---------------- Reports ----------------
elif section == "Reports":
    st.title("📦 Reports & Export")
    c1, c2, c3 = st.columns(3)
    with c1:
        company_f = st.selectbox("Company", ["All"] + sorted(usage["Company"].dropna().unique().tolist()))
    with c2:
        vehicle_f = st.selectbox("VehicleID", ["All"] + sorted(usage["VehicleID"].dropna().unique().tolist()))
    with c3:
        user_f = st.selectbox("User", ["All"] + sorted(usage["User"].dropna().unique().tolist()))

    df = usage.copy()
    if company_f!="All": df = df[df["Company"]==company_f]
    if vehicle_f!="All": df = df[df["VehicleID"]==vehicle_f]
    if user_f!="All": df = df[df["User"]==user_f]
    st.dataframe(df, use_container_width=True)

    csv = df.to_csv(index=False).encode("utf-8")
    st.download_button("Download CSV", data=csv, file_name="usage_export.csv", mime="text/csv")

    from io import BytesIO
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Usage", index=False)
    st.download_button("Download Excel", data=bio.getvalue(), file_name="usage_export.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------- Admin ----------------
elif section == "Admin" and role == "Admin":
    st.title("🛠️ Admin Panel")

    tab1, tab2 = st.tabs(["Add/Update Vehicle", "Add Usage Record"])

    with tab1:
        st.subheader("Add / Update Vehicle")
        with st.form("veh_form", clear_on_submit=False):
            v_id = st.text_input("VehicleID")
            plate = st.text_input("Plate No")
            company = st.text_input("Company")
            status = st.selectbox("Status", ["Available", "In Use"])
            odo = st.number_input("Odometer", min_value=0, step=1)
            last_service_odo = st.number_input("Last Service Odometer", min_value=0, step=1)
            last_service_date = st.date_input("Last Service Date", value=datetime.today())
            notes = st.text_area("Notes", value="")
            submit = st.form_submit_button("Save Vehicle")
        if submit:
            v = vehicles.copy()
            row = {"VehicleID":v_id, "PlateNo":plate, "Company":company, "Status":status,
                   "Odometer":odo, "LastServiceOdo":last_service_odo, "LastServiceDate":last_service_date, "Notes":notes}
            if (v["VehicleID"]==v_id).any():
                v.loc[v["VehicleID"]==v_id, :] = row
            else:
                v = pd.concat([v, pd.DataFrame([row])], ignore_index=True)
            save_sheet(v, "Vehicles")
            st.success("Vehicle saved.")
            st.cache_data.clear()
            st.rerun()

    with tab2:
        st.subheader("Add Usage")
        with st.form("usage_form", clear_on_submit=True):
            date = st.date_input("Date", value=datetime.today())
            user = st.text_input("User", value=auth["username"])
            company = st.text_input("Company", value=auth.get("company",""))
            vehicle_id = st.selectbox("Vehicle", options=vehicles["VehicleID"].tolist())
            if vehicle_id:
                plate = vehicles.set_index("VehicleID").loc[vehicle_id, "PlateNo"]
                current_odo = vehicles.set_index("VehicleID").loc[vehicle_id, "Odometer"]
            else:
                plate = ""
                current_odo = 0
            odo_start = st.number_input("Odo Start", min_value=0, step=1, value=int(current_odo))
            odo_end = st.number_input("Odo End", min_value=0, step=1, value=int(current_odo))
            purpose = st.text_input("Purpose")
            submit2 = st.form_submit_button("Save Usage")

        if submit2:
            negative_block = bool(get_setting("NEGATIVE_MILEAGE_BLOCK", True))
            dist = compute_distance(odo_start, odo_end)
            if negative_block and (dist is None or dist < 0):
                st.error("OdoEnd cannot be less than OdoStart.")
            else:
                new_id = (usage["LogID"].max() + 1) if not usage.empty else 1
                anomaly = flag_daily_anomaly(dist)
                new_row = {"LogID":new_id, "Date":date, "User":user, "Company":company, "VehicleID":vehicle_id,
                           "PlateNo":plate, "OdoStart":odo_start, "OdoEnd":odo_end, "Distance":dist,
                           "Purpose":purpose, "AnomalyFlag":anomaly, "AnomalyNote":"", "ApprovedBy":auth["username"]}
                u = pd.concat([usage, pd.DataFrame([new_row])], ignore_index=True)
                save_sheet(u, "UsageLog")

                v = vehicles.copy()
                v.loc[v["VehicleID"]==vehicle_id, "Odometer"] = odo_end
                v.loc[v["VehicleID"]==vehicle_id, "Status"] = "Available"
                save_sheet(v, "Vehicles")

                st.success(f"Usage saved. Distance={dist:.0f}. {'Anomaly flagged.' if anomaly else ''}")
                st.cache_data.clear()
                st.rerun()

st.caption("© 2025 Transport Record v2 • Excel backend • Streamlit UI")